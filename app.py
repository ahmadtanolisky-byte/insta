"""
InstaViral — Instagram Viral Post Scraper (Playwright Edition v5)
=================================================================
Uses saved cookies from login.py — no user login required.
Parser updated for Instagram's current xdt_api response format.

Setup (one time only):
  1. pip install flask playwright openpyxl
  2. playwright install chromium
  3. python login.py        ← do this once
  4. python app.py
  5. Open: http://127.0.0.1:5000
"""

from flask import Flask, request, jsonify, send_file
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io, json, re, time, random
from datetime import datetime, timezone
from pathlib import Path
from threading import Lock

COOKIES_FILE = "ig_cookies.json"

app  = Flask(__name__, static_folder='.', static_url_path='')
lock = Lock()
lock_acquired_at = None   # track when lock was grabbed
LOCK_TIMEOUT = 600        # auto-release after 10 minutes

# Live progress tracker
progress = {"count": 0, "phase": "", "active": False}


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    html = Path(__file__).parent / "index.html"
    if html.exists():
        return html.read_text(encoding="utf-8"), 200, {"Content-Type": "text/html"}
    return "<h3>index.html not found</h3>", 404


@app.route("/api/status")
def status():
    ok = Path(COOKIES_FILE).exists()
    return jsonify({"session_ok": ok, "account": "team-session" if ok else ""})


@app.route("/api/progress")
def get_progress():
    return jsonify(progress)


@app.route("/api/unlock", methods=["POST"])
def unlock():
    """Force-release the scrape lock — use if stuck."""
    global lock_acquired_at
    progress["active"] = False
    progress["phase"] = ""
    if lock.locked():
        try:
            lock.release()
        except RuntimeError:
            pass
    lock_acquired_at = None
    return jsonify({"ok": True, "message": "Lock released. You can scrape again."})


@app.route("/api/scrape", methods=["POST"])
def scrape():
    global lock_acquired_at

    if not Path(COOKIES_FILE).exists():
        return jsonify({"error": "No session found. Run 'python login.py' first."}), 503

    # Auto-release lock if it's been held for more than LOCK_TIMEOUT seconds
    if lock.locked() and lock_acquired_at is not None:
        if time.time() - lock_acquired_at > LOCK_TIMEOUT:
            try:
                lock.release()
            except RuntimeError:
                pass  # already released

    if not lock.acquire(blocking=False):
        elapsed = int(time.time() - lock_acquired_at) if lock_acquired_at else 0
        return jsonify({"error": f"Another scrape is running ({elapsed}s). Please wait and try again."}), 429

    lock_acquired_at = time.time()

    try:
        data     = request.json or {}
        username = data.get("username", "").strip().lstrip("@")
        limit    = int(data.get("limit", 0))
        if limit <= 0:
            limit = 99999  # 0 means "scrape everything"
        else:
            limit = min(max(limit, 1), 5000)
        sort_by  = data.get("sort_by", "likes")

        if not username:
            return jsonify({"error": "Username is required."}), 400

        progress["count"] = 0
        progress["phase"] = "Starting…"
        progress["active"] = True

        posts, error = get_posts(username, limit, sort_by)

        progress["active"] = False
        progress["phase"] = ""

        if error:
            return jsonify({"error": error}), 400

        return jsonify({"posts": posts, "count": len(posts), "username": username})
    finally:
        lock.release()


@app.route("/api/export", methods=["POST"])
def export():
    data     = request.json or {}
    posts    = data.get("posts", [])
    username = data.get("username", "account")
    sort_by  = data.get("sort_by", "likes")

    if not posts:
        return jsonify({"error": "No posts to export."}), 400

    buf      = build_excel(posts, username, sort_by)
    filename = f"instagram_viral_{username}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Core scraping ─────────────────────────────────────────────────────────────

def get_posts(username, display_limit, sort_by):
    all_posts  = []
    limit      = 99999   # always scrape full account; trim at the end

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-blink-features=AutomationControlled",
                    "--disable-dev-shm-usage",
                    "--disable-gpu",
                    "--disable-images",          # skip images — faster page load
                    "--blink-settings=imagesEnabled=false",
                ]
            )
            context = browser.new_context(
                viewport={"width": 1366, "height": 768},
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/122.0.0.0 Safari/537.36"
                ),
                locale="en-US",
                timezone_id="America/New_York",
                extra_http_headers={
                    "Accept-Language": "en-US,en;q=0.9",
                    "Accept-Encoding": "gzip, deflate, br",
                }
            )
            context.add_init_script(
                "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
                "Object.defineProperty(navigator,'languages',{get:()=>['en-US','en']});"
                "window.chrome={runtime:{},loadTimes:()=>{},csi:()=>{}};"
            )

            # Load saved cookies
            print("  → Loading session cookies...")
            with open(COOKIES_FILE) as f:
                cookies = json.load(f)
            context.add_cookies(cookies)
            print(f"  ✓ {len(cookies)} cookies loaded")

            # Read CSRF token once upfront
            csrf_token = ""
            for ck in cookies:
                if ck.get("name") == "csrftoken":
                    csrf_token = ck.get("value", "")
                    break

            # Intercept API responses to grab user_id and first batch
            intercepted = []
            def handle_response(response):
                try:
                    if "graphql/query" in response.url or "api/v1" in response.url:
                        try:
                            intercepted.append(response.json())
                        except Exception:
                            pass
                except Exception:
                    pass

            page = context.new_page()
            page.on("response", handle_response)

            # ── Load profile page ─────────────────────────────────────────
            print(f"  → Loading @{username}...")
            progress["phase"] = "Loading profile…"
            try:
                page.goto(
                    f"https://www.instagram.com/{username}/",
                    wait_until="domcontentloaded",
                    timeout=30000
                )
            except PlaywrightTimeout:
                browser.close()
                return None, "Timed out. Check your internet connection."

            time.sleep(random.uniform(2.5, 3.5))   # reduced from 3.5–5.0
            _dismiss_popups(page)

            content = page.content()
            print(f"  → Page loaded ({len(content)} chars)")

            if "/accounts/login" in page.url:
                browser.close()
                return None, "Session expired. Re-run 'python login.py' to refresh."

            if any(x in content for x in ["Page Not Found", "isn't available"]):
                browser.close()
                return None, f"Profile '@{username}' not found."

            if any(x in content for x in ["This Account is Private", "is_private\":true"]):
                browser.close()
                return None, "This account is private."

            # ── Extract user ID ───────────────────────────────────────────
            user_id = None
            for body in intercepted:
                uid = _find_user_id(body)
                if uid:
                    user_id = uid
                    break
            if not user_id:
                html = page.content()
                m = re.search(r'"pk"\s*:\s*"?(\d{5,})"?', html)
                if not m:
                    m = re.search(r'"id"\s*:\s*"(\d{5,})"', html)
                if m:
                    user_id = m.group(1)

            if user_id:
                print(f"  ✓ User ID: {user_id}")
            else:
                print("  ⚠ Could not extract user ID")
                browser.close()
                return None, "Could not find user ID. Try again."

            # ── Grab first batch from page load intercepts ────────────────
            batch, end_cursor = _extract_posts_and_cursor(intercepted)
            all_posts.extend(batch)
            intercepted.clear()
            progress["count"] = len(all_posts)
            print(f"  → Initial batch: {len(batch)} posts")

            # ── Helper: browser fetch (GET) ───────────────────────────────
            def _ig_get(url):
                raw = page.evaluate("""(args) => {
                    return fetch(args.url, {
                        headers: {
                            "X-IG-App-ID": "936619743392459",
                            "X-Requested-With": "XMLHttpRequest"
                        },
                        credentials: "include"
                    }).then(r => r.text().then(t => JSON.stringify({status: r.status, text: t})))
                      .catch(e => JSON.stringify({error: e.toString()}));
                }""", {"url": url})
                return json.loads(raw)

            # ── Helper: browser fetch (POST) ──────────────────────────────
            def _ig_post_req(url, body_str):
                result = page.evaluate("""(args) => {
                    return fetch(args.url, {
                        method: "POST",
                        headers: {
                            "X-IG-App-ID": "936619743392459",
                            "X-CSRFToken": args.csrf,
                            "X-Requested-With": "XMLHttpRequest",
                            "Content-Type": "application/x-www-form-urlencoded",
                            "Referer": "https://www.instagram.com/"
                        },
                        body: args.body,
                        credentials: "include"
                    }).then(r => r.text().then(t => ({status: r.status, text: t})))
                      .catch(e => ({error: e.toString()}));
                }""", {"url": url, "body": body_str, "csrf": csrf_token})
                return result

            # ── Phase 1: Feed API pagination (skips scroll entirely) ──────
            progress["phase"] = "Fetching posts via API…"
            print(f"  → Skipping scroll — going straight to API pagination...")

            max_id    = end_cursor
            api_fails = 0

            while api_fails < 8:
                base    = f"https://www.instagram.com/api/v1/feed/user/{user_id}/?count=200"
                api_url = base + (f"&max_id={max_id}" if max_id else "")

                try:
                    resp = _ig_get(api_url)
                    time.sleep(random.uniform(0.05, 0.1))   # minimal sleep

                    if "error" in resp:
                        raise ValueError(f"fetch error: {resp['error']}")
                    if resp.get("status") != 200:
                        raise ValueError(f"HTTP {resp.get('status')}")

                    body = json.loads(resp["text"])
                    batch, new_cursor = _extract_posts_from_api_v1(body)
                    if not batch:
                        batch, new_cursor = _extract_posts_and_cursor([body])

                    if batch:
                        all_posts.extend(batch)
                        max_id    = new_cursor
                        api_fails = 0
                        progress["count"] = len(all_posts)
                        print(f"  → Feed API: +{len(batch)} | total: {len(all_posts)} | more: {max_id is not None}")
                        if not max_id:
                            print("  → Reached end of feed.")
                            break
                    else:
                        api_fails += 1
                        print(f"  → Feed empty ({api_fails}/8)")
                        if api_fails >= 8:
                            # Feed API completely failed — fall back to scroll
                            if len(all_posts) == 0:
                                print("  → Feed API failed entirely — falling back to scroll...")
                                progress["phase"] = "Scrolling profile (fallback)…"
                                no_change = 0
                                while no_change < 4:
                                    page.evaluate("window.scrollTo({top:document.body.scrollHeight,behavior:'smooth'})")
                                    time.sleep(random.uniform(0.8, 1.2))
                                    scroll_batch, scroll_cursor = _extract_posts_and_cursor(intercepted)
                                    intercepted.clear()
                                    if scroll_batch:
                                        all_posts.extend(scroll_batch)
                                        end_cursor = scroll_cursor or end_cursor
                                        progress["count"] = len(all_posts)
                                        print(f"  → Scroll fallback: +{len(scroll_batch)} | total: {len(all_posts)}")
                                        no_change = 0
                                    else:
                                        no_change += 1
                                print(f"  → Scroll fallback done: {len(all_posts)} posts")
                            break
                        time.sleep(random.uniform(1.0, 2.0))

                except Exception as e:
                    api_fails += 1
                    print(f"  ⚠ Feed API error ({api_fails}/8): {e}")
                    time.sleep(random.uniform(1.0, 2.0))

            # ── Phase 2: Reels API (runs immediately after feed) ──────────
            if user_id and csrf_token:
                progress["phase"] = "Fetching Reels…"
                print(f"  → Fetching Reels for @{username}...")

                reel_max_id = None
                reel_fails  = 0
                reels_total = 0

                while reel_fails < 6:
                    body_str = f"target_user_id={user_id}&page_size=100"
                    if reel_max_id:
                        body_str += f"&max_id={reel_max_id}"

                    try:
                        probe = _ig_post_req(
                            "https://www.instagram.com/api/v1/clips/user/",
                            body_str
                        )
                        time.sleep(random.uniform(0.05, 0.1))   # minimal sleep

                        if "error" in probe:
                            raise ValueError(f"fetch error: {probe['error']}")
                        if probe.get("status") != 200:
                            raise ValueError(f"HTTP {probe.get('status')}: {probe.get('text','')[:80]}")

                        body_r = json.loads(probe["text"])
                        items  = body_r.get("items", [])

                        batch = []
                        for item in items:
                            node = item.get("media", item)
                            node["product_type"] = "clips"
                            if not node.get("shortcode") and node.get("code"):
                                node["shortcode"] = node["code"]
                            p = _parse_node(node)
                            if p:
                                batch.append(p)

                        more        = body_r.get("more_available", False)
                        reel_max_id = body_r.get("next_max_id") if more else None

                        if batch:
                            all_posts.extend(batch)
                            reels_total += len(batch)
                            reel_fails  = 0
                            progress["count"] = len(all_posts)
                            print(f"  → Reels: +{len(batch)} | total: {len(all_posts)} | more: {more}")
                        else:
                            reel_fails += 1
                            print(f"  → Reels empty ({reel_fails}/6)")

                        if not reel_max_id:
                            print(f"  → Reels done. Added: {reels_total}")
                            break

                    except Exception as re_err:
                        reel_fails += 1
                        print(f"  ⚠ Reels error ({reel_fails}/6): {re_err}")
                        time.sleep(random.uniform(1.0, 2.0))

            browser.close()

    except Exception as e:
        print(f"  ✗ Error: {e}")
        return None, f"Browser error: {str(e)[:200]}"

    if not all_posts:
        return None, "No posts found. The account may have no public posts, or the session expired — re-run 'python login.py'."

    all_posts = _dedup(all_posts)
    print(f"  → Done. Total unique posts: {len(all_posts)}")
    return _sort_and_trim(all_posts, sort_by, display_limit)



def _extract_posts_and_cursor(intercepted_list):
    """
    Parse Instagram's current xdt_api format.
    Response key: data.xdt_api__v1__feed__user_timeline_graphql_connection
    Each edge.node has: code, taken_at, like_count, comment_count,
                        caption.text, video_versions (if video)
    """
    posts     = []
    cursor    = None

    for body in intercepted_list:
        if not isinstance(body, dict):
            continue

        data = body.get("data", {})
        if not isinstance(data, dict):
            continue

        # ── New format: xdt_api__v1__feed__user_timeline_graphql_connection ──
        feed = data.get("xdt_api__v1__feed__user_timeline_graphql_connection")
        if feed and isinstance(feed, dict):
            edges     = feed.get("edges", [])
            page_info = feed.get("page_info", {})
            if page_info.get("has_next_page"):
                cursor = page_info.get("end_cursor")

            for edge in edges:
                node = edge.get("node", {})
                post = _parse_node(node)
                if post:
                    posts.append(post)
            continue

        # ── Old format: edge_owner_to_timeline_media (fallback) ──────────────
        user_data = data.get("user", {})
        if isinstance(user_data, dict):
            media = user_data.get("edge_owner_to_timeline_media", {})
            if media:
                page_info = media.get("page_info", {})
                if page_info.get("has_next_page"):
                    cursor = page_info.get("end_cursor")
                for edge in media.get("edges", []):
                    node = edge.get("node", {})
                    post = _parse_node_old(node)
                    if post:
                        posts.append(post)
                continue

        # ── Recursive search as last resort ───────────────────────────────────
        found = _deep_find_posts(body)
        posts.extend(found)

    return _dedup(posts), cursor


def _extract_reels(intercepted_list):
    """
    Parse Reels from the /reels/ tab.
    Instagram uses these response keys for the Reels feed:
      - xdt_api__v1__clips__home_timeline_connection_v2
      - xdt_api__v1__clips__home_timeline_connection
      - items[] array (older fallback)
    Each item/node is parsed with _parse_node (product_type='clips' marks it as Reel).
    """
    posts  = []
    cursor = None

    for body in intercepted_list:
        if not isinstance(body, dict):
            continue

        data = body.get("data", {})
        if not isinstance(data, dict):
            # Some reels responses are flat {items: [...]}
            items = body.get("items", [])
            for item in items:
                node = item.get("media", item)
                # Force product_type so _parse_node labels it correctly
                if "product_type" not in node:
                    node["product_type"] = "clips"
                p = _parse_node(node)
                if p:
                    posts.append(p)
            continue

        # New clips connection key (v2 and v1)
        for clips_key in (
            "xdt_api__v1__clips__home_timeline_connection_v2",
            "xdt_api__v1__clips__home_timeline_connection",
        ):
            feed = data.get(clips_key)
            if feed and isinstance(feed, dict):
                page_info = feed.get("page_info", {})
                if page_info.get("has_next_page"):
                    cursor = page_info.get("end_cursor")
                for edge in feed.get("edges", []):
                    node = edge.get("node", {})
                    # Reels edges may wrap the media under "media" key
                    media_node = node.get("media", node)
                    if "product_type" not in media_node:
                        media_node["product_type"] = "clips"
                    p = _parse_node(media_node)
                    if p:
                        posts.append(p)
                break

        # Fallback: items[] array at root or inside data
        for items_key in ("items",):
            items = data.get(items_key, [])
            for item in items:
                node = item.get("media", item)
                if "product_type" not in node:
                    node["product_type"] = "clips"
                p = _parse_node(node)
                if p:
                    posts.append(p)

    return _dedup(posts), cursor


def _extract_view_count(intercepted_list, shortcode):
    """
    Extract view/play count from intercepted API responses on a post page.
    Instagram sends the media detail via graphql/query when visiting /p/<code>/.
    """
    for item in intercepted_list:
        body = item.get("body", {}) if isinstance(item, dict) else {}
        if not isinstance(body, dict):
            continue

        # Search recursively for play_count / view_count in the response
        count = _find_view_count_in_obj(body)
        if count:
            return count
    return 0


def _find_view_count_in_obj(obj, depth=0):
    """Recursively search a dict/list for a non-null play_count or view_count."""
    if depth > 12 or not obj:
        return 0
    if isinstance(obj, dict):
        for key in ("play_count", "video_view_count", "view_count"):
            v = obj.get(key)
            if v and isinstance(v, (int, float)) and v > 0:
                return int(v)
        for v in obj.values():
            result = _find_view_count_in_obj(v, depth + 1)
            if result:
                return result
    elif isinstance(obj, list):
        for item in obj:
            result = _find_view_count_in_obj(item, depth + 1)
            if result:
                return result
    return 0


def _scrape_view_count_from_html(html):
    """
    Fallback: parse view count from page HTML using known patterns.
    Instagram embeds structured data and meta tags with view counts.
    """
    import re
    # Pattern 1: JSON-LD or script with play_count / video_view_count
    for pattern in [
        r'"play_count"\s*:\s*(\d+)',
        r'"video_view_count"\s*:\s*(\d+)',
        r'"view_count"\s*:\s*(\d+)',
        r'"playCount"\s*:\s*(\d+)',
    ]:
        m = re.search(pattern, html)
        if m:
            val = int(m.group(1))
            if val > 0:
                return val
    return 0


def _parse_node(node):
    """Parse a post node from the new xdt_api format."""
    if not isinstance(node, dict):
        return None

    # Shortcode field varies by API endpoint:
    # - Feed API:  "code"
    # - Reels API: "code" or derivable from "pk" via base64
    sc = node.get("code") or node.get("shortcode") or ""

    # Fallback: derive shortcode from numeric pk using Instagram's base64 encoding
    if not sc:
        pk = node.get("pk") or node.get("id") or ""
        if pk:
            try:
                pk_int = int(str(pk).split("_")[0])
                alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_"
                result = []
                while pk_int > 0:
                    result.append(alphabet[pk_int % 64])
                    pk_int //= 64
                sc = "".join(reversed(result))
            except Exception:
                sc = str(pk)  # use raw pk as fallback URL key

    if not sc:
        return None

    # media_type: 1=image, 2=video, 8=carousel
    # product_type: "clips" = Reel
    media_type   = node.get("media_type", 1)
    product_type = node.get("product_type", "")
    is_reel      = product_type == "clips"
    is_video     = media_type == 2 or bool(node.get("video_versions")) or is_reel

    likes    = node.get("like_count", 0) or 0
    comments = node.get("comment_count", 0) or 0
    views    = (node.get("play_count") or node.get("ig_play_count") or
               node.get("video_view_count") or node.get("view_count") or 0)

    # Caption is nested: caption.text
    cap = node.get("caption") or {}
    if isinstance(cap, dict):
        caption = cap.get("text", "")
    else:
        caption = str(cap)

    ts       = node.get("taken_at", 0) or 0
    date_str = datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d") if ts else ""

    return {
        "url":       f"https://www.instagram.com/p/{sc}/",
        "shortcode": sc,
        "likes":     int(likes),
        "comments":  int(comments),
        "views":     int(views),
        "type":      "Reel" if is_reel else ("Video" if is_video else ("Carousel" if media_type == 8 else "Image")),
        "caption":   _trim(caption),
        "date":      date_str,
        "is_video":  is_video,
    }


def _parse_node_old(node):
    """Parse a post node from the old GraphQL edge format (fallback)."""
    if not isinstance(node, dict):
        return None

    sc = node.get("shortcode") or node.get("code", "")
    if not sc:
        return None

    is_video = node.get("is_video", False)
    likes    = (node.get("edge_liked_by") or {}).get("count", 0) or node.get("like_count", 0) or 0
    comments = (node.get("edge_media_to_comment") or {}).get("count", 0) or node.get("comment_count", 0) or 0
    views    = node.get("video_view_count") or node.get("view_count") or 0

    cap_edges = (node.get("edge_media_to_caption") or {}).get("edges", [])
    if cap_edges:
        caption = cap_edges[0].get("node", {}).get("text", "")
    else:
        c = node.get("caption") or ""
        caption = c.get("text", "") if isinstance(c, dict) else str(c)

    ts       = node.get("taken_at_timestamp") or node.get("taken_at") or 0
    date_str = datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d") if ts else ""

    return {
        "url":       f"https://www.instagram.com/p/{sc}/",
        "shortcode": sc,
        "likes":     int(likes),
        "comments":  int(comments),
        "views":     int(views),
        "type":      "Video" if is_video else "Image",
        "caption":   _trim(caption),
        "date":      date_str,
        "is_video":  is_video,
    }


def _deep_find_posts(obj, depth=0):
    """Recursive fallback search for any post arrays."""
    if depth > 10 or not isinstance(obj, dict):
        return []

    posts = []

    # New format key
    feed = obj.get("xdt_api__v1__feed__user_timeline_graphql_connection")
    if feed and isinstance(feed, dict):
        for edge in feed.get("edges", []):
            p = _parse_node(edge.get("node", {}))
            if p:
                posts.append(p)
        return posts

    # Old format key
    media = obj.get("edge_owner_to_timeline_media")
    if media and isinstance(media, dict):
        for edge in media.get("edges", []):
            p = _parse_node_old(edge.get("node", {}))
            if p:
                posts.append(p)
        return posts

    for v in obj.values():
        if isinstance(v, dict):
            posts.extend(_deep_find_posts(v, depth + 1))

    return posts


def _find_user_id(obj, depth=0):
    """Recursively search API response for the account's numeric user ID."""
    if depth > 8 or not obj:
        return None
    if isinstance(obj, dict):
        # Common keys that hold the profile user ID
        for key in ("pk", "pk_id", "id"):
            v = obj.get(key)
            if v and isinstance(v, (str, int)):
                s = str(v)
                if s.isdigit() and len(s) >= 6:
                    return s
        for v in obj.values():
            result = _find_user_id(v, depth + 1)
            if result:
                return result
    elif isinstance(obj, list):
        for item in obj[:5]:  # only check first few items
            result = _find_user_id(item, depth + 1)
            if result:
                return result
    return None


def _dedup(posts):
    """Deduplicate by shortcode, keeping the entry with the highest view count."""
    seen = {}  # shortcode -> index in out
    out = []
    for p in posts:
        sc = p.get("shortcode", "") if isinstance(p, dict) else ""
        if not sc:
            out.append(p)
            continue
        if sc not in seen:
            seen[sc] = len(out)
            out.append(p)
        else:
            # Prefer whichever version has a non-zero view count
            idx = seen[sc]
            existing_views = out[idx].get("views", 0) or 0
            new_views = p.get("views", 0) or 0
            if new_views > existing_views:
                out[idx] = p
    return out


def _extract_posts_from_api_v1(body):
    """
    Parse Instagram's /api/v1/feed/user/<username>/username/ response.
    Response shape: { items: [...], next_max_id: "...", more_available: true }
    Each item is a media node compatible with _parse_node / _parse_node_old.
    """
    posts  = []
    cursor = None

    if not isinstance(body, dict):
        return posts, cursor

    # Pagination cursor
    if body.get("more_available") and body.get("next_max_id"):
        cursor = body["next_max_id"]

    items = body.get("items", [])
    for item in items:
        # v1 feed wraps media directly as the item
        post = _parse_node(item)
        if not post:
            post = _parse_node_old(item)
        if post:
            posts.append(post)

    return _dedup(posts), cursor



def _sort_and_trim(posts, sort_by, limit):
    key = {
        "likes":    lambda x: x["likes"],
        "comments": lambda x: x["comments"],
        "views":    lambda x: x["views"],
        "mixed":    lambda x: x["likes"] * 0.4 + x["comments"] * 0.4 + x["views"] * 0.2,
    }.get(sort_by, lambda x: x["likes"])
    posts.sort(key=key, reverse=True)
    # limit=99999 means "all" — don't actually slice to 99999
    if limit >= 99999:
        return posts, None
    return posts[:limit], None


def _dismiss_popups(page):
    for sel in [
        "button:has-text('Accept All')",
        "button:has-text('Allow all cookies')",
        "button:has-text('Not Now')",
        "button:has-text('Save Info')",
        "[aria-label='Close']",
    ]:
        try:
            btn = page.query_selector(sel)
            if btn:
                btn.click()
                time.sleep(0.4)
        except Exception:
            pass


def _trim(text, n=150):
    if not text:
        return ""
    text = str(text).strip()
    return (text[:n] + "…") if len(text) > n else text


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(posts, username, sort_by):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Viral Posts"

    hdr_fill  = PatternFill("solid", fgColor="1a1a2e")
    alt_fill  = PatternFill("solid", fgColor="16213e")
    norm_fill = PatternFill("solid", fgColor="0f3460")
    border    = Border(**{s: Side(style="thin", color="e94560")
                          for s in ("left","right","top","bottom")})

    ws.merge_cells("A1:H1")
    t           = ws["A1"]
    t.value     = (f"Instagram Viral Posts — @{username}  |  "
                   f"Sorted by: {sort_by.upper()}  |  "
                   f"{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    t.font      = Font(color="e94560", bold=True, size=13)
    t.fill      = hdr_fill
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(["#","Post URL","Type","Likes ❤️",
                              "Comments 💬","Views 👁️","Date 📅","Caption"], 1):
        c           = ws.cell(row=2, column=col, value=h)
        c.font      = Font(color="e94560", bold=True, size=11)
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
    ws.row_dimensions[2].height = 22

    for i, p in enumerate(posts, 1):
        row  = i + 2
        fill = alt_fill if i % 2 == 0 else norm_fill
        ws.row_dimensions[row].height = 18
        for col, val in enumerate([i, p["url"], p["type"], p["likes"],
                                    p["comments"], p["views"], p["date"], p["caption"]], 1):
            c           = ws.cell(row=row, column=col, value=val)
            c.fill      = fill
            c.border    = border
            c.alignment = Alignment(
                horizontal="left" if col == 8 else "center",
                vertical="center", wrap_text=(col == 8))
            if col == 2:
                c.hyperlink = val
                c.font = Font(color="a8d8ea", size=10, underline="single")
            else:
                c.font = Font(color="FFFFFF", size=10)

    for i, w in enumerate([5,45,10,14,14,14,12,55], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Start ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n" + "=" * 55)
    print("  InstaViral — Instagram Scraper (Playwright v5)")
    print("=" * 55)

    if not Path(COOKIES_FILE).exists():
        print("\n  ⚠  No session found!")
        print("  Run 'python login.py' first.\n")
    else:
        print("\n  ✓  Session ready — no user login needed")
        print("  Open in browser → http://127.0.0.1:5000\n")

    import os
    app.run(debug=False, port=int(os.environ.get("PORT", 5000)), host="0.0.0.0")